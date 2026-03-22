"""
Build Nanodegree-Integrated-Capstone.xlsx — advanced Excel capstone tying together
marketing metrics, P&L/margins, forecasting, stickiness, unit economics, MLB salaries,
and world cities. Run from repo root: python scripts/build_capstone_workbook.py
"""
from __future__ import annotations

import csv
import re
from collections import Counter
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
OUT_DIR = ROOT / "excel-capstone"
OUT_PATH = OUT_DIR / "Nanodegree-Integrated-Capstone.xlsx"

HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
INPUT_FILL = PatternFill("solid", fgColor="FFF2CC")
CALC_FILL = PatternFill("solid", fgColor="E2EFDA")
THIN = Side(style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def money_cell(
    ws,
    row: int,
    col: int,
    value: float | int | None = None,
    *,
    formula: str | None = None,
    is_input: bool = False,
):
    c = ws.cell(row=row, column=col, value=formula if formula is not None else value)
    c.number_format = '"$"#,##0.00'
    c.border = BORDER
    if is_input:
        c.fill = INPUT_FILL


def pct_cell(ws, row: int, col: int, value=None, formula: str | None = None):
    c = ws.cell(row=row, column=col, value=formula if formula else value)
    c.number_format = "0.00%"
    c.border = BORDER


def header_row(ws, row: int, labels: list[str], start_col: int = 1):
    for i, lab in enumerate(labels):
        c = ws.cell(row=row, column=start_col + i, value=lab)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center", wrap_text=True)
        c.border = BORDER


def parse_money(s: str) -> float:
    if s is None:
        return 0.0
    s = str(s).strip().strip('"')
    s = re.sub(r"[^\d.\-]", "", s)
    return float(s) if s else 0.0


def load_marketing_raw():
    path = ROOT / "band-marketingmetric-quizzes-a.csv"
    rows = []
    with path.open(newline="", encoding="utf-8") as f:
        r = csv.DictReader(f)
        for row in r:
            rows.append(
                {
                    "campaign": row["Ad campaign"].strip(),
                    "impr": int(row["Impressions"]),
                    "clicks": int(row["Clicks"]),
                    "spend": parse_money(row[" Total spent on ad campaign "]),
                    "leads": int(row["Total leads"]),
                }
            )
    return rows


def load_pl_smoothie():
    """Use margin quiz + pl statement for revenue, units, and major costs."""
    path = ROOT / "smoothierock-marginquizzes.csv"
    with path.open(newline="", encoding="utf-8") as f:
        lines = list(csv.reader(f))
    revenue = parse_money(lines[1][1])
    units = 60895
    for row in lines:
        if row and row[0] and "units sold" in row[0].lower():
            raw = str(row[1]).replace(",", "").strip() if len(row) > 1 else ""
            if raw.isdigit():
                units = int(raw)
            break
    return revenue, units


def load_pl_lines():
    path = ROOT / "smoothierock-plstatementquiz.csv"
    items = []
    with path.open(newline="", encoding="utf-8") as f:
        r = csv.reader(f)
        for row in r:
            if not row or not row[0].strip():
                continue
            label = row[0].strip().strip('"').strip()
            v1 = parse_money(row[1]) if len(row) > 1 else 0.0
            items.append((label, v1))
    return items


def load_stickiness():
    path = ROOT / "band-stickiness-quiz.csv"
    out = []
    with path.open(newline="", encoding="utf-8") as f:
        r = csv.reader(f)
        next(r)  # header row with empty first cell
        for row in r:
            if len(row) < 3 or not row[1].strip():
                continue
            month = row[0].strip().strip(",")
            dau = int(row[1])
            mau = int(row[2])
            out.append((month, dau, mau))
    return out


def load_mlb():
    path = ROOT / "albb-salaries-2003.csv"
    rows = []
    with path.open(newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f, delimiter="\t")
        for row in reader:
            sal = parse_money(row["Salary"])
            rows.append(
                {
                    "team": row["Team"].strip(),
                    "last": row["Last"].strip(),
                    "first": row["First"].strip(),
                    "salary": sal,
                    "position": row["Position"].strip(),
                }
            )
    return rows


def load_world_cities():
    path = ROOT / "worldcities.csv"
    rows = []
    with path.open(newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            rows.append((row["City"].strip(), row["Country"].strip()))
    return rows


def load_monthly_unit_econ():
    """CPA quiz (costs + leads) and quizzes-b (paid customers)."""
    cpa_path = ROOT / "band-cpa-quiz.csv"
    cust_path = ROOT / "band-marketingmetric-quizzes-b.csv"
    months = ["January", "February", "March"]
    marketing = []
    salaries = []
    overhead = []
    leads = []
    with cpa_path.open(newline="", encoding="utf-8") as f:
        r = list(csv.reader(f))
    marketing = [parse_money(r[1][i]) for i in range(1, 4)]
    salaries = [parse_money(r[2][i]) for i in range(1, 4)]
    overhead = [parse_money(r[3][i]) for i in range(1, 4)]
    leads = [int(r[5][i]) for i in range(1, 4)]
    with cust_path.open(newline="", encoding="utf-8") as f:
        r = list(csv.reader(f))
    customers = [int(r[4][i]) for i in range(1, 4)]
    return list(zip(months, marketing, salaries, overhead, leads, customers))


def build():
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    wb = Workbook()

    # --- Start Here ---
    ws0 = wb.active
    ws0.title = "Start_Here"
    ws0.column_dimensions["A"].width = 110
    intro = """NANODEGREE INTEGRATED EXCEL CAPSTONE

This workbook ties together themes from the Data Analyst Nanodegree practice sets: marketing KPIs (CTR, CPC, CPL), P&L and gross margin, sales capacity forecasting, stickiness (DAU/MAU), monthly unit economics and cost per lead, exploratory analysis on MLB salaries, and geography lookups.

HOW TO USE IT
1. Work through sheets in the suggested order below. Yellow cells are inputs you may change for what-if analysis; green-tinted areas are calculation zones (verify or extend formulas).
2. On Marketing_Campaigns, confirm CTR, CPC, and CPL match the definitions in your course materials.
3. On PnL_Margins, tie revenue and costs to gross margin and (optional) contribution margin using the labeled lines.
4. On Sales_Capacity_Model, adjust Model_Start_Date and ramp months to see how hiring timing affects first-year bookings.
5. Insert your own charts (recommended): CTR by campaign; stickiness over time; salary mix by position; top countries from Cities_Summary.
6. Dashboard pulls live KPIs from other sheets — refresh/recalc if you change inputs (F9).

SHEET MAP (suggested order)
Start_Here (you are here) → Marketing_Campaigns → PnL_Margins → Monthly_Unit_Economics → Stickiness → Sales_Capacity_Model → MLB_Salaries → MLB_by_Team → World_Cities_Raw → Cities_Summary → Lookup_Lab → Dashboard

DATA SOURCES (CSV files in repo root)
band-marketingmetric-quizzes-a.csv, smoothierock-*.csv, band-cpa-quiz.csv, band-marketingmetric-quizzes-b.csv, band-stickiness-quiz.csv, practice-quiz-salesforecast-topdown-data-csvfile.csv, albb-salaries-2003.csv, worldcities.csv
"""
    ws0["A1"] = intro
    ws0["A1"].alignment = Alignment(wrap_text=True, vertical="top")
    ws0["A1"].font = Font(size=11)
    ws0.row_dimensions[1].height = 480

    # --- Marketing ---
    mkt = wb.create_sheet("Marketing_Campaigns")
    mkt.column_dimensions["A"].width = 18
    for c in range(2, 9):
        mkt.column_dimensions[get_column_letter(c)].width = 14
    hdr = [
        "Ad campaign",
        "Impressions",
        "Clicks",
        "Spend",
        "Leads",
        "CTR",
        "CPC",
        "CPL",
    ]
    header_row(mkt, 1, hdr)
    data = load_marketing_raw()
    for i, d in enumerate(data, start=2):
        mkt.cell(row=i, column=1, value=d["campaign"]).border = BORDER
        mkt.cell(row=i, column=2, value=d["impr"]).border = BORDER
        mkt.cell(row=i, column=3, value=d["clicks"]).border = BORDER
        money_cell(mkt, i, 4, d["spend"])
        mkt.cell(row=i, column=5, value=d["leads"]).border = BORDER
        pct_cell(mkt, i, 6, formula=f"=IF(B{i}=0,\"\",C{i}/B{i})")
        money_cell(mkt, i, 7, formula=f"=IF(C{i}=0,\"\",D{i}/C{i})")
        money_cell(mkt, i, 8, formula=f"=IF(E{i}=0,\"\",D{i}/E{i})")
    last = 1 + len(data)
    mkt.cell(row=last + 1, column=1, value="TOTAL / AVG").font = Font(bold=True)
    mkt.cell(row=last + 1, column=2, value=f"=SUM(B2:B{last})").border = BORDER
    mkt.cell(row=last + 1, column=3, value=f"=SUM(C2:C{last})").border = BORDER
    money_cell(mkt, last + 1, 4, formula=f"=SUM(D2:D{last})")
    mkt.cell(row=last + 1, column=5, value=f"=SUM(E2:E{last})").border = BORDER
    pct_cell(mkt, last + 1, 6, formula=f"=IF(B{last+1}=0,\"\",C{last+1}/B{last+1})")
    money_cell(mkt, last + 1, 7, formula=f"=IF(C{last+1}=0,\"\",D{last+1}/C{last+1})")
    money_cell(mkt, last + 1, 8, formula=f"=IF(E{last+1}=0,\"\",D{last+1}/E{last+1})")

    # Chart: CTR by campaign
    chart_ctr = BarChart()
    chart_ctr.type = "col"
    chart_ctr.title = "CTR by campaign"
    chart_ctr.y_axis.title = "CTR"
    chart_ctr.x_axis.title = "Campaign"
    data_ref = Reference(mkt, min_col=6, min_row=1, max_row=last, max_col=6)
    cats = Reference(mkt, min_col=1, min_row=2, max_row=last)
    chart_ctr.add_data(data_ref, titles_from_data=True)
    chart_ctr.set_categories(cats)
    chart_ctr.height = 8
    chart_ctr.width = 18
    mkt.add_chart(chart_ctr, "J2")

    # --- P&L ---
    pl = wb.create_sheet("PnL_Margins")
    pl.column_dimensions["A"].width = 36
    pl.column_dimensions["B"].width = 16
    rev, units = load_pl_smoothie()
    pl_items = load_pl_lines()
    expense_rows: list[tuple[int, str]] = []
    header_row(pl, 1, ["Line item", "Amount ($)"], 1)
    r = 2
    pl.cell(row=r, column=1, value="Total Revenue").border = BORDER
    money_cell(pl, r, 2, rev, is_input=True)
    r += 1
    for label, amt in pl_items:
        if "Revenue" in label and amt == rev:
            continue
        pl.cell(row=r, column=1, value=label).border = BORDER
        money_cell(pl, r, 2, amt)
        expense_rows.append((r, label))
        r += 1
    pl.cell(row=r, column=1, value="Units sold").border = BORDER
    pl.cell(row=r, column=2, value=units).border = BORDER
    pl.cell(row=r, column=1).fill = INPUT_FILL
    pl.cell(row=r, column=2).fill = INPUT_FILL
    units_row_idx = r
    last_expense_row = units_row_idx - 1
    cogs_row = next(
        (row for row, lab in expense_rows if "fruits" in lab.lower() and "vegetable" in lab.lower()),
        expense_rows[0][0],
    )
    r += 2
    pl.cell(row=r, column=1, value="Gross profit (Revenue − COGS fruits/veg)").font = Font(bold=True)
    money_cell(pl, r, 2, formula=f"=B2-B{cogs_row}")
    pl.cell(row=r, column=2).fill = CALC_FILL
    gp_row = r
    r += 1
    pl.cell(row=r, column=1, value="Gross margin %").font = Font(bold=True)
    pct_cell(pl, r, 2, formula=f"=IF(B2=0,\"\",B{gp_row}/B2)")
    pl.cell(row=r, column=2).fill = CALC_FILL
    gm_row = r
    r += 1
    pl.cell(row=r, column=1, value="Operating profit (Revenue − all costs above units)").font = Font(bold=True)
    money_cell(pl, r, 2, formula=f"=B2-SUM(B3:B{last_expense_row})")
    pl.cell(row=r, column=2).fill = CALC_FILL
    r += 2
    pl.cell(row=r, column=1, value="Contribution margin / unit (GP ÷ units)").font = Font(bold=True)
    money_cell(
        pl,
        r,
        2,
        formula=f"=IF(B{units_row_idx}=0,\"\",B{gp_row}/B{units_row_idx})",
    )
    pl.cell(row=r, column=2).fill = CALC_FILL

    # --- Monthly unit economics ---
    mu = wb.create_sheet("Monthly_Unit_Economics")
    header_row(
        mu,
        1,
        [
            "Month",
            "Marketing $",
            "S&M Salaries $",
            "Salary OH $",
            "Leads",
            "Paid customers",
            "New paid (Δ)",
            "Cost / lead",
            "Mktg $ / new paid",
        ],
    )
    monthly = load_monthly_unit_econ()
    for i, (mo, mk, sal, oh, leads, cust) in enumerate(monthly, start=2):
        mu.cell(row=i, column=1, value=mo).border = BORDER
        money_cell(mu, i, 2, mk, is_input=True)
        money_cell(mu, i, 3, sal, is_input=True)
        money_cell(mu, i, 4, oh, is_input=True)
        mu.cell(row=i, column=5, value=leads).border = BORDER
        mu.cell(row=i, column=6, value=cust).border = BORDER
        if i == 2:
            mu.cell(row=i, column=7, value="—").border = BORDER
        else:
            money_cell(mu, i, 7, formula=f"=F{i}-F{i-1}")
        money_cell(mu, i, 8, formula=f"=IF(E{i}=0,\"\",B{i}/E{i})")
        money_cell(mu, i, 9, formula=f"=IF(OR(G{i}=\"—\",G{i}=0),\"\",B{i}/G{i})")
    mu.column_dimensions["A"].width = 12
    for c in range(2, 10):
        mu.column_dimensions[get_column_letter(c)].width = 16

    # --- Stickiness ---
    st = wb.create_sheet("Stickiness")
    header_row(st, 1, ["Month", "DAU", "MAU", "Stickiness (DAU/MAU)", "Course check"])
    stick = load_stickiness()
    for i, (mo, dau, mau) in enumerate(stick, start=2):
        st.cell(row=i, column=1, value=mo).border = BORDER
        st.cell(row=i, column=2, value=dau).border = BORDER
        st.cell(row=i, column=3, value=mau).border = BORDER
        pct_cell(st, i, 4, formula=f"=IF(C{i}=0,\"\",B{i}/C{i})")
        # Column E: reference values from CSV for self-check
        st.cell(row=i, column=5, value=round(dau / mau, 9)).border = BORDER
    st.column_dimensions["A"].width = 12
    lr = 1 + len(stick)
    lc = LineChart()
    lc.title = "Stickiness (DAU/MAU)"
    lc.y_axis.title = "Ratio"
    lc.add_data(Reference(st, min_col=4, min_row=1, max_row=lr, max_col=4), titles_from_data=True)
    lc.set_categories(Reference(st, min_col=1, min_row=2, max_row=lr))
    lc.height = 7
    lc.width = 16
    st.add_chart(lc, "G2")

    # --- Sales capacity (from practice forecast assumptions) ---
    sm = wb.create_sheet("Sales_Capacity_Model")
    sm.column_dimensions["A"].width = 34
    sm.column_dimensions["B"].width = 18
    r = 1
    header_row(sm, r, ["Assumption", "Value"], 1)
    r = 2
    labels = [
        ("Model_Start_Date (editable)", "2025-01-01"),
        ("Opportunities closed / rep / year", 53),
        ("Avg PPU / month ($)", 250),
        ("Avg units per opportunity", 56),
        ("Avg contract months / opportunity", 24),
        ("Avg opportunity size — booking ($)", 336000),
        ("Annual booking capacity / rep ($)", "=B3*B7"),
        ("Seller ramp (months to full productivity)", 4),
    ]
    for lab, val in labels:
        sm.cell(row=r, column=1, value=lab).border = BORDER
        if isinstance(val, str) and val.startswith("="):
            sm.cell(row=r, column=2, value=val).border = BORDER
            sm.cell(row=r, column=2).fill = CALC_FILL
            sm.cell(row=r, column=2).number_format = '"$"#,##0.00'
        elif lab.startswith("Model_Start"):
            sm.cell(row=r, column=2, value=val).border = BORDER
            sm.cell(row=r, column=2).number_format = "yyyy-mm-dd"
            sm.cell(row=r, column=2).fill = INPUT_FILL
        else:
            c = sm.cell(row=r, column=2, value=val)
            c.border = BORDER
            if "$" in lab or "booking" in lab.lower() or "PPU" in lab:
                c.number_format = '"$"#,##0.00' if "PPU" in lab or "booking" in lab.lower() else "0"
                if "booking" in lab.lower() or "PPU" in lab:
                    c.fill = INPUT_FILL
            else:
                c.fill = INPUT_FILL
        r += 1
    r += 1
    header_row(sm, r, ["Sales employee", "Hire date", "Ramp end (EDATE)", "Yr-1 fraction (YEARFRAC)", "Yr-1 bookings ($)"], 1)
    start_r = r + 1
    offsets = [30, 60, 90]
    for j, off in enumerate(offsets):
        row = start_r + j
        sm.cell(row=row, column=1, value=f"Sales person {j+1}").border = BORDER
        sm.cell(row=row, column=2, value=f"=$B$2+{off}").border = BORDER
        sm.cell(row=row, column=2).number_format = "yyyy-mm-dd"
        sm.cell(row=row, column=3, value=f"=EDATE(B{row},$B$9)").border = BORDER
        sm.cell(row=row, column=3).number_format = "yyyy-mm-dd"
        sm.cell(row=row, column=4, value=f"=MIN(1,MAX(0,YEARFRAC(B{row},DATE(2025,12,31),1)))").border = BORDER
        sm.cell(row=row, column=4).number_format = "0.00%"
        # Bookings: scale annual capacity by fraction of year after ramp, simplified = YEARFRAC(ramp_end, EOY)*annual * min(1,...)
        sm.cell(row=row, column=5, value=f"=MIN($B$7,$B$7*MAX(0,YEARFRAC(C{row},DATE(2025,12,31),1)))").border = BORDER
        sm.cell(row=row, column=5).number_format = '"$"#,##0.00'
        sm.cell(row=row, column=5).fill = CALC_FILL
    last_rep = start_r + len(offsets) - 1
    sales_total_row = last_rep + 1
    sm.cell(row=sales_total_row, column=1, value="Total Yr-1 bookings (3 reps)").font = Font(bold=True)
    money_cell(sm, sales_total_row, 5, formula=f"=SUM(E{start_r}:E{last_rep})")
    sm.cell(row=sales_total_row, column=5).fill = CALC_FILL
    r = last_rep + 3
    sm.cell(row=r, column=1, value="Note: Formulas are a simplified ramp model for practice; refine with monthly grids for production forecasts.").font = Font(italic=True, size=10)

    # --- MLB ---
    mlb = wb.create_sheet("MLB_Salaries")
    mlb_rows = load_mlb()
    header_row(mlb, 1, ["Team", "Last", "First", "Salary", "Position"], 1)
    for i, row in enumerate(mlb_rows, start=2):
        mlb.cell(row=i, column=1, value=row["team"])
        mlb.cell(row=i, column=2, value=row["last"])
        mlb.cell(row=i, column=3, value=row["first"])
        c = mlb.cell(row=i, column=4, value=row["salary"])
        c.number_format = '"$"#,##0.00'
        mlb.cell(row=i, column=5, value=row["position"])
    for c in range(1, 6):
        mlb.column_dimensions[get_column_letter(c)].width = 18

    # Position totals for pie chart (aggregated in Python for clarity)
    by_pos = Counter(r["position"] for r in mlb_rows)
    by_team_total = {}
    for r in mlb_rows:
        by_team_total[r["team"]] = by_team_total.get(r["team"], 0) + r["salary"]

    mt = wb.create_sheet("MLB_by_Team")
    header_row(mt, 1, ["Team", "Total payroll", "Players"], 1)
    teams_sorted = sorted(by_team_total.keys())
    for i, team in enumerate(teams_sorted, start=2):
        mt.cell(row=i, column=1, value=team).border = BORDER
        money_cell(mt, i, 2, formula=f'=SUMIFS(MLB_Salaries!$D:$D,MLB_Salaries!$A:$A,A{i})')
        mt.cell(row=i, column=3, value=f'=COUNTIF(MLB_Salaries!$A:$A,A{i})').border = BORDER
    last_t = 1 + len(teams_sorted)
    mt.cell(row=last_t + 1, column=1, value="League total").font = Font(bold=True)
    money_cell(mt, last_t + 1, 2, formula=f"=SUM(B2:B{last_t})")

    pos = wb.create_sheet("MLB_by_Position")
    header_row(pos, 1, ["Position", "Headcount", "Total salary"], 1)
    for i, (pname, cnt) in enumerate(sorted(by_pos.items()), start=2):
        pos.cell(row=i, column=1, value=pname).border = BORDER
        pos.cell(row=i, column=2, value=cnt).border = BORDER
        money_cell(pos, i, 3, sum(r["salary"] for r in mlb_rows if r["position"] == pname))
    pie = PieChart()
    pie.title = "Share of payroll by position"
    pdata = Reference(pos, min_col=3, min_row=1, max_row=1 + len(by_pos), max_col=3)
    plabs = Reference(pos, min_col=1, min_row=2, max_row=1 + len(by_pos))
    pie.add_data(pdata, titles_from_data=True)
    pie.set_categories(plabs)
    pie.height = 10
    pie.width = 12
    pos.add_chart(pie, "E2")

    # --- Cities ---
    cities = load_world_cities()
    wc = wb.create_sheet("World_Cities_Raw")
    header_row(wc, 1, ["City", "Country"], 1)
    for i, (city, country) in enumerate(cities, start=2):
        wc.cell(row=i, column=1, value=city)
        wc.cell(row=i, column=2, value=country)
    wc.column_dimensions["A"].width = 28
    wc.column_dimensions["B"].width = 28

    ctr = Counter(c for _, c in cities)
    top = ctr.most_common(40)
    cs = wb.create_sheet("Cities_Summary")
    header_row(cs, 1, ["Country", "City count"], 1)
    for i, (country, n) in enumerate(top, start=2):
        cs.cell(row=i, column=1, value=country).border = BORDER
        cs.cell(row=i, column=2, value=n).border = BORDER

    # --- Lookup lab ---
    ll = wb.create_sheet("Lookup_Lab")
    ll["A1"] = "Use INDEX/MATCH or XLOOKUP (Excel 365) against World_Cities_Raw to return the country for a city."
    ll["A1"].font = Font(bold=True)
    ll["A3"] = "Sample city (change me)"
    ll["B3"] = "Salzburg"
    ll["B3"].fill = INPUT_FILL
    ll["A4"] = "Country (your formula)"
    ll["B4"] = '=IFERROR(INDEX(World_Cities_Raw!$B:$B,MATCH(B3,World_Cities_Raw!$A:$A,0)),"not found")'
    ll["A6"] = "Count of cities in United States of America (try COUNTIF on World_Cities_Raw col B)"
    ll["B6"] = '=COUNTIF(World_Cities_Raw!$B:$B,"United States of America")'
    for row in range(3, 7):
        ll.cell(row=row, column=1).border = BORDER
        ll.cell(row=row, column=2).border = BORDER
    ll.column_dimensions["A"].width = 55
    ll.column_dimensions["B"].width = 45

    # --- Dashboard ---
    mkt_total_row = last + 1
    march_stick_row = next((i for i, s in enumerate(stick, start=2) if str(s[0]).lower().startswith("march")), 4)
    mlb_total_row = last_t + 1
    cs_last_row = 1 + len(top)
    top_country_formula = (
        f"=INDEX(Cities_Summary!$A$2:$A${cs_last_row},"
        f"MATCH(MAX(Cities_Summary!$B$2:$B${cs_last_row}),Cities_Summary!$B$2:$B${cs_last_row},0))"
    )

    dash = wb.create_sheet("Dashboard")
    dash.column_dimensions["A"].width = 38
    dash.column_dimensions["B"].width = 22
    header_row(dash, 1, ["KPI", "Value", "Source"], 1)
    rows_d = [
        ("Blended CTR (all campaigns)", f"=Marketing_Campaigns!F{mkt_total_row}", "Marketing_Campaigns"),
        ("Total ad spend", f"=Marketing_Campaigns!D{mkt_total_row}", "Marketing_Campaigns"),
        ("Gross margin %", f"=PnL_Margins!B{gm_row}", "PnL_Margins"),
        ("March stickiness (DAU/MAU)", f"=Stickiness!D{march_stick_row}", "Stickiness"),
        ("Total Yr-1 sales bookings (3 reps)", f"=Sales_Capacity_Model!E{sales_total_row}", "Sales_Capacity_Model"),
        ("League payroll (MLB sample)", f"=MLB_by_Team!B{mlb_total_row}", "MLB_by_Team"),
        ("Top country by city count", top_country_formula, "Cities_Summary"),
    ]
    for i, (kpi, formula, src) in enumerate(rows_d, start=2):
        dash.cell(row=i, column=1, value=kpi).border = BORDER
        dash.cell(row=i, column=2, value=formula).border = BORDER
        dash.cell(row=i, column=3, value=src).border = BORDER
        if "margin" in kpi.lower() or "stickiness" in kpi.lower() or "CTR" in kpi:
            dash.cell(row=i, column=2).number_format = "0.00%"
        elif "payroll" in kpi.lower() or "spend" in kpi.lower() or "bookings" in kpi.lower():
            dash.cell(row=i, column=2).number_format = '"$"#,##0.00'

    wb.save(OUT_PATH)
    print(f"Wrote {OUT_PATH}")


if __name__ == "__main__":
    build()
